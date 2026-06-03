"""
Runner determinístico para task_code.

O código da task deve definir uma função `run(from_date, to_date, ctx)` que
recebe as datas do período e um TaskContext, executa a lógica e retorna
o(s) token(s) dos artifacts gerados (str ou list[str]).

Execução via exec() com namespace controlado — apenas bibliotecas explicitamente
aprovadas ficam disponíveis. Sem acesso a os, subprocess, sys, open, etc.
"""

import builtins
import traceback
from datetime import date, timedelta

from .context import TaskContext


# ── Bibliotecas aprovadas para o namespace das tasks ─────────────────────────

def _build_globals(ctx: TaskContext, from_date: str, to_date: str) -> dict:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import matplotlib.dates as mdates
    import numpy as np
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import date as _date, datetime as _datetime, timedelta as _timedelta
    import time as _time
    import math as _math
    import json as _json
    import re as _re_mod
    from collections import Counter as _Counter, defaultdict as _defaultdict

    # Módulos stdlib leves pré-injetados — o LLM costuma tentar importá-los;
    # ao disponibilizá-los diretamente evitamos o loop import→strip→NameError→reimport.
    _ALLOWED = {"time": _time, "math": _math, "json": _json, "re": _re_mod}

    # Pré-carrega _strptime para que datetime.strptime() funcione no sandbox.
    # Python importa _strptime lazily na primeira chamada a strptime.
    import _strptime  # noqa: F401

    # __import__ precisa existir no dict de builtins ou o Python levanta KeyError
    # internamente. Módulos da whitelist são permitidos; módulos já presentes em
    # sys.modules também são permitidos (imports internos lazy do Python, como _strptime);
    # qualquer outro é bloqueado com mensagem clara.
    import sys as _sys
    def _blocked_import(name, *args, **kwargs):
        if name in _ALLOWED:
            return _ALLOWED[name]
        if name in _sys.modules:
            return _sys.modules[name]
        raise ImportError(
            f"Import de '{name}' não é permitido dentro do run(). "
            "Use as variáveis já disponíveis no namespace: "
            "pd, np, plt, date, datetime, timedelta, time, math, json, re, openpyxl, ctx, etc."
        )

    safe_builtins = {
        name: getattr(builtins, name)
        for name in (
            "abs", "all", "any", "bool", "dict", "divmod", "enumerate",
            "filter", "float", "format", "frozenset", "getattr", "hasattr",
            "hash", "int", "isinstance", "issubclass", "iter", "len", "list",
            "map", "max", "min", "next", "object", "print", "range", "repr",
            "reversed", "round", "set", "slice", "sorted", "str", "sum",
            "tuple", "type", "zip", "True", "False", "None",
            "ValueError", "TypeError", "KeyError", "IndexError",
            "RuntimeError", "StopIteration", "Exception",
        )
    }
    safe_builtins["__import__"] = _blocked_import

    # Converte para objetos date para que .strftime(), .year, etc. funcionem.
    # str(date_obj) ainda retorna YYYY-MM-DD, então f'/api?from={from_date}' continua OK.
    from_date_obj = _date.fromisoformat(from_date) if isinstance(from_date, str) else from_date
    to_date_obj   = _date.fromisoformat(to_date)   if isinstance(to_date, str)   else to_date

    return {
        "__builtins__": safe_builtins,
        # datas do período — objetos date (suportam .strftime(), .year etc.)
        # str(from_date) == 'YYYY-MM-DD', então f-strings da API continuam funcionando
        "from_date": from_date_obj,
        "to_date":   to_date_obj,
        # datetime — não usar import dentro do run(), já está disponível
        "date":      _date,
        "datetime":  _datetime,
        "timedelta": _timedelta,
        # contexto com api() e save_*()
        "ctx": ctx,
        # matplotlib
        "plt":     plt,
        "mticker": mticker,
        "mdates":  mdates,
        # numpy / pandas
        "np":  np,
        "pd":  pd,
        # openpyxl
        "openpyxl":          openpyxl,
        "Font":              Font,
        "PatternFill":       PatternFill,
        "Alignment":         Alignment,
        "Border":            Border,
        "Side":              Side,
        "get_column_letter": get_column_letter,
        # stdlib leve — pré-injetados para evitar erros quando o LLM usa sem import
        "time":        _time,
        "math":        _math,
        "json":        _json,
        "re":          _re_mod,
        "Counter":     _Counter,
        "defaultdict": _defaultdict,
    }


# ── Executor ──────────────────────────────────────────────────────────────────

class TaskCodeError(Exception):
    """Erro de validação ou execução do task_code."""


def run_task_code(code: str, from_date: str, to_date: str, session_id: str, user_id: str | None = None, is_test: bool = False, notify_enabled: bool = True, task_name: str = ""):
    """Compila e executa task_code.

    Retorna list[str] de tokens em modo normal.
    Retorna (list[str], TaskContext) em modo teste para inspeção de alertas capturados.
    Lança TaskCodeError com mensagem amigável em caso de falha.
    """
    ctx = TaskContext(session_id, user_id=user_id, is_test=is_test, notify_enabled=notify_enabled, task_name=task_name)
    g   = _build_globals(ctx, from_date, to_date)

    # 1a. Rejeita imports antes de compilar, exceto módulos já pré-injetados no namespace.
    # Módulos permitidos (pré-injetados em _build_globals): time, math, json, re.
    import re as _re
    _ALLOWED_MODS = {"time", "math", "json", "re"}

    def _is_blocked_import(ln: str) -> bool:
        m = _re.match(r'\s*(import\s+(\S+)|from\s+(\S+)\s+import)', ln)
        if not m:
            return "__import__" in ln and not ln.lstrip().startswith("#")
        mod = (m.group(2) or m.group(3) or "").split(".")[0]
        return mod not in _ALLOWED_MODS

    _import_lines = [ln.strip() for ln in code.splitlines() if _is_blocked_import(ln)]
    if _import_lines:
        raise TaskCodeError(
            "O task_code contém imports, que são bloqueados no sandbox:\n"
            + "\n".join(f"  {ln}" for ln in _import_lines)
            + "\n\nRemova essas linhas — as seguintes variáveis já estão disponíveis:\n"
            "  pd, np, plt, mticker, date, datetime, timedelta, time, math, json, re,\n"
            "  Counter, defaultdict, openpyxl, Font, PatternFill, Alignment, Border, Side,\n"
            "  get_column_letter, ctx, from_date, to_date"
        )

    # 1b. Compila — detecta erros de sintaxe antes de executar
    try:
        compiled = compile(code, "<task_code>", "exec")
    except SyntaxError as e:
        raise TaskCodeError(f"Erro de sintaxe no código:\n{e}") from e

    # 2. Executa o bloco — registra a função `run` no namespace
    try:
        exec(compiled, g)
    except Exception as e:
        raise TaskCodeError(
            f"Erro ao carregar o código:\n{traceback.format_exc(limit=8)}"
        ) from e

    # 3. Valida que `run` foi definido
    if "run" not in g or not callable(g["run"]):
        raise TaskCodeError(
            "O código deve definir uma função `run(from_date, to_date, ctx)`."
        )

    # 4. Chama run()
    try:
        result = g["run"](g["from_date"], g["to_date"], ctx)
    except Exception as e:
        raise TaskCodeError(
            f"Erro durante a execução de run():\n{traceback.format_exc(limit=12)}"
        ) from e

    # 4b. Auto-notificação: se notify estava habilitado e o código não chamou ctx.notify(),
    # dispara automaticamente com o nome da tarefa — sem depender do LLM escrever ctx.notify().
    if notify_enabled and not is_test and not ctx._notify_called:
        ctx.notify(ctx._task_name or "Tarefa executada com sucesso")

    # 5. Coleta tokens — aceita str, list ou None (usa ctx.tokens())
    if isinstance(result, str):
        tokens = [result]
    elif isinstance(result, (list, tuple)):
        tokens = [str(t) for t in result]
    else:
        tokens = ctx.tokens()

    if not tokens:
        # Monitores retornam string de status sem artifact — aceito se houver alertas ou string de status
        if is_test and (ctx.test_alerts() or isinstance(result, str)):
            tokens = [result] if isinstance(result, str) else ["(sem artifact — monitor)"]
        else:
            raise TaskCodeError(
                "run() não retornou nenhum token de artifact. "
                "Use ctx.save_chart(), ctx.save_excel() ou ctx.save_pdf()."
            )

    return (tokens, ctx) if is_test else tokens


def default_test_range() -> tuple[str, str]:
    """Período padrão para testes: últimos 7 dias."""
    end   = date.today()
    start = end - timedelta(days=6)
    return start.isoformat(), end.isoformat()
